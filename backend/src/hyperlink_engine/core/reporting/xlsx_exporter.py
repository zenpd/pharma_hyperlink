"""W8.4 — XLSX anomaly and link-record exporter.

Produces a multi-sheet openpyxl workbook with:

* **Links** sheet   — one row per link (mirrors the CSV report with status
  colour-coding: green = ok, yellow = suspicious/unverified, red = broken).
* **Anomalies** sheet — one row per anomaly, colour-coded by severity.
* **Summary** sheet  — document-level KPIs (link count, broken count,
  blocker anomalies, readiness score).
* **Pivot** sheet    — anomaly count by kind × severity (for management views).

All cell formatting uses ``openpyxl``; the file is written as a real .xlsx
(zip-based OOXML) that opens in Microsoft Excel and LibreOffice Calc.

Usage::

    from hyperlink_engine.core.reporting.xlsx_exporter import write_xlsx_report
    write_xlsx_report(
        path=Path("output/report.xlsx"),
        link_records=records,
        anomaly_summary=summary,
        readiness_result=readiness,
    )
"""

from __future__ import annotations

from collections import defaultdict
from pathlib import Path
from typing import Iterable, Sequence

from hyperlink_engine.config.logging_setup import get_logger
from hyperlink_engine.core.reporting.readiness_score import ReadinessResult
from hyperlink_engine.core.validation.anomaly_detector import DossierAnomalySummary
from hyperlink_engine.models import (
    Anomaly,
    AnomalyKind,
    AnomalySeverity,
    LinkRecord,
    LinkStatus,
)

_log = get_logger("reporting.xlsx_exporter")

# ── openpyxl is a soft dependency — raise a clear error if absent ─────────
try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    _OPENPYXL_AVAILABLE = True
except ImportError:  # pragma: no cover
    _OPENPYXL_AVAILABLE = False


# ─────────────────────────────────────────────────────────────────────────────
# Colour palette
# ─────────────────────────────────────────────────────────────────────────────

_FILL_GREEN = "C6EFCE"   # ok / A-grade
_FILL_YELLOW = "FFEB9C"  # warning / suspicious
_FILL_RED = "FFC7CE"     # broken / blocker
_FILL_BLUE = "DAEEF3"    # header row
_FILL_GREY = "F2F2F2"    # alternating row background

_FONT_HEADER = dict(bold=True, color="000000")
_FONT_RED = dict(bold=True, color="9C0006")
_FONT_YELLOW = dict(bold=True, color="9C5700")
_FONT_GREEN = dict(bold=True, color="276221")


def _fill(hex_color: str) -> "PatternFill":
    return PatternFill(fill_type="solid", fgColor=hex_color)


def _link_fill(status: LinkStatus) -> "PatternFill":
    if status == LinkStatus.BROKEN:
        return _fill(_FILL_RED)
    if status in (LinkStatus.SUSPICIOUS, LinkStatus.UNVERIFIED):
        return _fill(_FILL_YELLOW)
    return _fill(_FILL_GREEN)


def _anomaly_fill(severity: AnomalySeverity) -> "PatternFill":
    if severity == AnomalySeverity.BLOCKER:
        return _fill(_FILL_RED)
    if severity == AnomalySeverity.WARNING:
        return _fill(_FILL_YELLOW)
    return _fill(_FILL_GREEN)


# ─────────────────────────────────────────────────────────────────────────────
# Sheet builders
# ─────────────────────────────────────────────────────────────────────────────


def _autofit(ws: "openpyxl.worksheet.worksheet.Worksheet") -> None:
    """Set column widths based on max content length (approximate)."""
    for col_cells in ws.columns:
        max_len = max(
            (len(str(cell.value)) if cell.value is not None else 0 for cell in col_cells),
            default=0,
        )
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(
            max(max_len + 2, 10), 60
        )


def _write_header(ws: "openpyxl.worksheet.worksheet.Worksheet", columns: list[str]) -> None:
    ws.append(columns)
    for cell in ws[1]:
        cell.font = Font(**_FONT_HEADER)
        cell.fill = _fill(_FILL_BLUE)
        cell.alignment = Alignment(wrap_text=False)
    ws.freeze_panes = ws["A2"]


def _build_links_sheet(
    ws: "openpyxl.worksheet.worksheet.Worksheet",
    link_records: Iterable[LinkRecord],
) -> None:
    headers = [
        "source_doc", "link_text", "link_location", "target_doc",
        "target_anchor", "status", "confidence", "error_msg",
    ]
    _write_header(ws, headers)

    for idx, rec in enumerate(link_records, start=2):
        row = [
            rec.source_doc,
            rec.link_text,
            rec.link_location_descriptor,
            rec.target_doc or "",
            rec.target_anchor or "",
            rec.status.value,
            round(rec.confidence, 3),
            rec.error_msg or "",
        ]
        ws.append(row)
        row_fill = _link_fill(rec.status)
        for cell in ws[idx]:
            cell.fill = row_fill

    _autofit(ws)


def _build_anomalies_sheet(
    ws: "openpyxl.worksheet.worksheet.Worksheet",
    anomalies: Iterable[Anomaly],
) -> None:
    headers = [
        "document", "kind", "severity", "text", "suggested_fix", "confidence",
    ]
    _write_header(ws, headers)

    for idx, anomaly in enumerate(anomalies, start=2):
        row = [
            anomaly.document,
            anomaly.kind.value,
            anomaly.severity.value,
            anomaly.text,
            anomaly.suggested_fix or "",
            round(anomaly.confidence, 3),
        ]
        ws.append(row)
        row_fill = _anomaly_fill(anomaly.severity)
        for cell in ws[idx]:
            cell.fill = row_fill

    _autofit(ws)


def _build_summary_sheet(
    ws: "openpyxl.worksheet.worksheet.Worksheet",
    readiness: ReadinessResult | None,
) -> None:
    headers = [
        "metric", "value",
    ]
    _write_header(ws, headers)

    if readiness is None:
        ws.append(["No readiness data available", ""])
        return

    rows = [
        ("Overall Score", f"{readiness.overall_score:.1f}/100"),
        ("Grade", readiness.grade),
        ("Submission Ready", "YES" if readiness.is_submission_ready else "NO"),
        ("Documents Processed", readiness.document_count),
        ("Total Links", readiness.total_links),
        ("Broken Links", readiness.broken_links),
        ("Orphaned References", readiness.orphaned_refs),
        ("Style Violations", readiness.style_violations),
        ("Blocker Anomalies", readiness.blocker_anomalies),
        ("Warning Anomalies", readiness.warning_anomalies),
        ("Broken Link Rate", f"{readiness.broken_rate*100:.2f}%"),
    ]
    for metric, value in rows:
        ws.append([metric, value])

    # Add module breakdown if available
    if readiness.module_scores:
        ws.append([])
        ws.append(["Module Breakdown", ""])
        ws.append(["module", "score", "broken_links", "blockers", "warnings", "docs"])
        for ms in sorted(readiness.module_scores, key=lambda x: x.score):
            ws.append([
                ms.module, round(ms.score, 1),
                ms.broken_links, ms.blocker_anomalies,
                ms.warning_anomalies, ms.document_count,
            ])

    _autofit(ws)


def _build_pivot_sheet(
    ws: "openpyxl.worksheet.worksheet.Worksheet",
    anomalies: list[Anomaly],
) -> None:
    """Anomaly count pivot: rows = kind, columns = severity."""
    severities = [AnomalySeverity.BLOCKER, AnomalySeverity.WARNING, AnomalySeverity.INFO]
    kinds = list(AnomalyKind)

    headers = ["Anomaly Kind"] + [s.value.upper() for s in severities] + ["TOTAL"]
    _write_header(ws, headers)

    counts: dict[AnomalyKind, dict[AnomalySeverity, int]] = defaultdict(
        lambda: defaultdict(int)
    )
    for a in anomalies:
        counts[a.kind][a.severity] += 1

    for kind in kinds:
        row = [kind.value]
        total = 0
        for sev in severities:
            count = counts[kind].get(sev, 0)
            row.append(count)
            total += count
        row.append(total)
        ws.append(row)

    # Totals row
    ws.append([])
    total_row = ["TOTAL"]
    grand_total = 0
    for sev in severities:
        col_total = sum(counts[k].get(sev, 0) for k in kinds)
        total_row.append(col_total)
        grand_total += col_total
    total_row.append(grand_total)
    ws.append(total_row)
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)

    _autofit(ws)


# ─────────────────────────────────────────────────────────────────────────────
# Public API
# ─────────────────────────────────────────────────────────────────────────────


def write_xlsx_report(
    *,
    path: Path,
    link_records: Sequence[LinkRecord] | None = None,
    anomaly_summary: DossierAnomalySummary | None = None,
    readiness_result: ReadinessResult | None = None,
) -> Path:
    """Write a multi-sheet XLSX report to ``path``.

    Parameters
    ----------
    path:
        Output file path (.xlsx).
    link_records:
        All link records across the dossier batch.
    anomaly_summary:
        Aggregated anomaly data (from
        :func:`validation.anomaly_detector.aggregate_anomaly_reports`).
    readiness_result:
        Readiness score result (from
        :func:`reporting.readiness_score.compute_readiness_score`).

    Returns
    -------
    Path
        The written file path.
    """
    if not _OPENPYXL_AVAILABLE:  # pragma: no cover
        raise ImportError(
            "openpyxl is required for XLSX export. "
            "Install it with: pip install openpyxl"
        )

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()

    # Sheet 1 — Links
    ws_links = wb.active
    ws_links.title = "Links"
    _build_links_sheet(ws_links, link_records or [])

    # Sheet 2 — Anomalies
    ws_anom = wb.create_sheet("Anomalies")
    all_anomalies = anomaly_summary.all_anomalies if anomaly_summary else []
    _build_anomalies_sheet(ws_anom, all_anomalies)

    # Sheet 3 — Summary
    ws_summary = wb.create_sheet("Summary")
    _build_summary_sheet(ws_summary, readiness_result)

    # Sheet 4 — Pivot
    ws_pivot = wb.create_sheet("Pivot")
    _build_pivot_sheet(ws_pivot, all_anomalies)

    wb.save(str(path))
    _log.info(
        "xlsx_report_written",
        path=str(path),
        links=len(link_records or []),
        anomalies=len(all_anomalies),
    )
    return path
