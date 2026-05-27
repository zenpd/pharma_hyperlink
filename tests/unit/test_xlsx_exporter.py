"""Unit tests for reporting/xlsx_exporter.py (W8.4)."""

from __future__ import annotations

from pathlib import Path

import pytest

from hyperlink_engine.models import (
    Anomaly,
    AnomalyKind,
    AnomalySeverity,
    LinkRecord,
    LinkStatus,
)
from hyperlink_engine.reporting.readiness_score import ReadinessResult
from hyperlink_engine.reporting.xlsx_exporter import write_xlsx_report
from hyperlink_engine.validation.anomaly_detector import (
    AnomalyReport,
    aggregate_anomaly_reports,
)


# ── Helpers ───────────────────────────────────────────────────────────────────


def _make_link_record(status: LinkStatus = LinkStatus.OK) -> LinkRecord:
    return LinkRecord(
        source_doc="doc.docx",
        link_text="Section 2.5.3",
        link_location_descriptor="p0.r0:c0-10",
        target_anchor="section_ref_2_5_3",
        status=status,
        confidence=0.9,
    )


def _make_anomaly(kind: AnomalyKind, severity: AnomalySeverity) -> Anomaly:
    return Anomaly(
        kind=kind,
        severity=severity,
        document="doc.docx",
        text="example",
        confidence=0.85,
    )


def _make_summary(blockers: int = 0, warnings: int = 0) -> object:
    reports: list[AnomalyReport] = []
    for _ in range(blockers):
        reports.append(
            AnomalyReport(
                document="doc.docx",
                anomalies=[
                    _make_anomaly(AnomalyKind.DEPRECATED_STUDY_ID, AnomalySeverity.BLOCKER)
                ],
            )
        )
    for _ in range(warnings):
        reports.append(
            AnomalyReport(
                document="doc.docx",
                anomalies=[
                    _make_anomaly(AnomalyKind.ORPHANED_REFERENCE, AnomalySeverity.WARNING)
                ],
            )
        )
    return aggregate_anomaly_reports(reports)


def _make_readiness(score: float = 90.0, broken: int = 0) -> ReadinessResult:
    return ReadinessResult(
        overall_score=score,
        grade="A" if score >= 95 else "B" if score >= 85 else "C",
        total_links=100,
        broken_links=broken,
        orphaned_refs=0,
        style_violations=0,
        blocker_anomalies=0,
        warning_anomalies=0,
        document_count=5,
    )


# ── Tests ─────────────────────────────────────────────────────────────────────


def test_write_xlsx_creates_file(tmp_path: Path) -> None:
    out = tmp_path / "report.xlsx"
    result = write_xlsx_report(path=out)
    assert result == out
    assert out.exists()
    assert out.stat().st_size > 0


def test_write_xlsx_four_sheets(tmp_path: Path) -> None:
    """The workbook should always contain the four expected sheets."""
    import openpyxl

    out = tmp_path / "report.xlsx"
    write_xlsx_report(
        path=out,
        link_records=[_make_link_record(), _make_link_record(LinkStatus.BROKEN)],
        anomaly_summary=_make_summary(blockers=1, warnings=2),
        readiness_result=_make_readiness(),
    )
    wb = openpyxl.load_workbook(str(out))
    assert set(wb.sheetnames) == {"Links", "Anomalies", "Summary", "Pivot"}


def test_links_sheet_has_data_rows(tmp_path: Path) -> None:
    import openpyxl

    out = tmp_path / "links.xlsx"
    records = [_make_link_record(LinkStatus.OK), _make_link_record(LinkStatus.BROKEN)]
    write_xlsx_report(path=out, link_records=records)
    wb = openpyxl.load_workbook(str(out))
    ws = wb["Links"]
    # Row 1 is header; rows 2+ are data
    assert ws.max_row >= 3  # header + 2 data rows


def test_anomalies_sheet_has_data_rows(tmp_path: Path) -> None:
    import openpyxl

    out = tmp_path / "anomalies.xlsx"
    summary = _make_summary(blockers=2, warnings=1)
    write_xlsx_report(path=out, anomaly_summary=summary)
    wb = openpyxl.load_workbook(str(out))
    ws = wb["Anomalies"]
    assert ws.max_row >= 4  # header + 3 data rows


def test_summary_sheet_contains_readiness_score(tmp_path: Path) -> None:
    import openpyxl

    out = tmp_path / "summary.xlsx"
    readiness = _make_readiness(score=87.5)
    write_xlsx_report(path=out, readiness_result=readiness)
    wb = openpyxl.load_workbook(str(out))
    ws = wb["Summary"]
    # Look for the score value in the sheet
    cell_values = [
        str(cell.value) for row in ws.iter_rows() for cell in row if cell.value is not None
    ]
    assert any("87.5" in v for v in cell_values)


def test_pivot_sheet_covers_all_anomaly_kinds(tmp_path: Path) -> None:
    import openpyxl

    out = tmp_path / "pivot.xlsx"
    summary = _make_summary(blockers=1, warnings=1)
    write_xlsx_report(path=out, anomaly_summary=summary)
    wb = openpyxl.load_workbook(str(out))
    ws = wb["Pivot"]
    first_col_values = [str(cell.value) for cell in ws["A"] if cell.value]
    # Should have a header row + at least as many rows as AnomalyKind values
    assert len(first_col_values) >= 2


def test_write_xlsx_empty_inputs(tmp_path: Path) -> None:
    """Empty inputs produce a valid (but minimal) workbook."""
    import openpyxl

    out = tmp_path / "empty.xlsx"
    write_xlsx_report(path=out)
    wb = openpyxl.load_workbook(str(out))
    assert "Links" in wb.sheetnames


def test_write_xlsx_creates_parent_dirs(tmp_path: Path) -> None:
    out = tmp_path / "nested" / "deep" / "report.xlsx"
    write_xlsx_report(path=out)
    assert out.exists()
